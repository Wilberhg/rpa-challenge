from playwright.sync_api import sync_playwright, expect
from openpyxl import load_workbook

wb = load_workbook("./rpa-challenge/challenge.xlsx")
ws = wb.active

with sync_playwright() as p:
    browser = p.chromium.launch(headless=False)
    page = browser.new_page()
    page.goto("https://rpachallenge.com/")
    page.locator("button").click()
    
    for row in ws.iter_rows(min_row=2, max_row=11, max_col=7, values_only=True):
        first_name,	last_name, company_name, role_company, address, email, phone_number = row
        page.locator('input[ng-reflect-name="labelFirstName"]').evaluate("(campo, texto) => campo.value = texto", first_name)
        page.locator('input[ng-reflect-name="labelRole"]').evaluate("(campo, texto) => campo.value = texto", role_company)
        page.locator('input[ng-reflect-name="labelAddress"]').evaluate("(campo, texto) => campo.value = texto", address)
        page.locator('input[ng-reflect-name="labelCompanyName"]').evaluate("(campo, texto) => campo.value = texto", company_name)
        page.locator('input[ng-reflect-name="labelPhone"]').evaluate("(campo, texto) => campo.value = texto", str(phone_number))
        page.locator('input[ng-reflect-name="labelLastName"]').evaluate("(campo, texto) => campo.value = texto", last_name)
        page.locator('input[ng-reflect-name="labelEmail"]').evaluate("(campo, texto) => campo.value = texto", email)
        page.locator('input[type="submit"]').click()
    ...
    browser.close()

