from selenium import webdriver
from selenium.webdriver.common.by import By
from openpyxl import load_workbook

wb = load_workbook("./rpa-challenge/challenge.xlsx")
ws = wb.active

driver = webdriver.Chrome()
driver.implicitly_wait(30)
driver.get("https://rpachallenge.com/")
driver.find_element(By.TAG_NAME, "button").click()
for row in ws.iter_rows(min_row=2, max_row=11, max_col=7, values_only=True):
    first_name,	last_name, company_name, role_company, address, email, phone_number = row
    driver.find_element(By.CSS_SELECTOR, 'input[ng-reflect-name="labelFirstName"]').send_keys(first_name)
    driver.find_element(By.CSS_SELECTOR, 'input[ng-reflect-name="labelRole"]').send_keys(role_company)
    driver.find_element(By.CSS_SELECTOR, 'input[ng-reflect-name="labelAddress"]').send_keys(address)
    driver.find_element(By.CSS_SELECTOR, 'input[ng-reflect-name="labelCompanyName"]').send_keys(company_name)
    driver.find_element(By.CSS_SELECTOR, 'input[ng-reflect-name="labelPhone"]').send_keys(phone_number)
    driver.find_element(By.CSS_SELECTOR, 'input[ng-reflect-name="labelLastName"]').send_keys(last_name)
    driver.find_element(By.CSS_SELECTOR, 'input[ng-reflect-name="labelEmail"]').send_keys(email)
    driver.find_element(By.CSS_SELECTOR, 'input[type="submit"]').click()
input()
driver.quit()