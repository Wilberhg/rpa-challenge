from selenium import webdriver
from selenium.webdriver.common.by import By
from openpyxl import load_workbook

wb = load_workbook("./rpa-challenge/challenge.xlsx")
ws = wb.active

driver = webdriver.Chrome()
driver.implicitly_wait(30)
driver.get("https://rpachallenge.com/")
driver.find_element(By.TAG_NAME, "button").click()
for row in ws.iter_rows(min_row=2, max_row=11, max_col=7, values_only=True):
    first_name,	last_name, company_name, role_company, address, email, phone_number = row
    driver.execute_script("arguments[0].value = arguments[1]", driver.find_element(By.CSS_SELECTOR, 'input[ng-reflect-name="labelFirstName"]'), first_name)
    driver.execute_script("arguments[0].value = arguments[1]", driver.find_element(By.CSS_SELECTOR, 'input[ng-reflect-name="labelRole"]'), role_company)
    driver.execute_script("arguments[0].value = arguments[1]", driver.find_element(By.CSS_SELECTOR, 'input[ng-reflect-name="labelAddress"]'), address)
    driver.execute_script("arguments[0].value = arguments[1]", driver.find_element(By.CSS_SELECTOR, 'input[ng-reflect-name="labelCompanyName"]'), company_name)
    driver.execute_script("arguments[0].value = arguments[1]", driver.find_element(By.CSS_SELECTOR, 'input[ng-reflect-name="labelPhone"]'), phone_number)
    driver.execute_script("arguments[0].value = arguments[1]", driver.find_element(By.CSS_SELECTOR, 'input[ng-reflect-name="labelLastName"]'), last_name)
    driver.execute_script("arguments[0].value = arguments[1]", driver.find_element(By.CSS_SELECTOR, 'input[ng-reflect-name="labelEmail"]'), email)
    driver.execute_script("arguments[0].click()", driver.find_element(By.CSS_SELECTOR, 'input[type="submit"]'))
input()
driver.quit()