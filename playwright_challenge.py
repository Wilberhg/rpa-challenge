from playwright.sync_api import sync_playwright, expect
from openpyxl import load_workbook

wb = load_workbook("./rpa-challenge/challenge.xlsx")
ws = wb.active

with sync_playwright() as p:
    browser = p.chromium.launch(headless=False)
    page = browser.new_page()
    page.goto("https://rpachallenge.com/")
    page.locator("button").click()
    
    for row in ws.iter_rows(min_row=2, max_row=11, max_col=7, values_only=True):
        first_name,	last_name, company_name, role_company, address, email, phone_number = row
        page.locator('input[ng-reflect-name="labelFirstName"]').fill(first_name)
        page.locator('input[ng-reflect-name="labelRole"]').fill(role_company)
        page.locator('input[ng-reflect-name="labelAddress"]').fill(address)
        page.locator('input[ng-reflect-name="labelCompanyName"]').fill(company_name)
        page.locator('input[ng-reflect-name="labelPhone"]').fill(str(phone_number))
        page.locator('input[ng-reflect-name="labelLastName"]').fill(last_name)
        page.locator('input[ng-reflect-name="labelEmail"]').fill(email)
        page.locator('input[type="submit"]').click()
    input()
    browser.close()

